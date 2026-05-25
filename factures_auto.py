#!/usr/bin/env python3
"""
Invoice Automation Script
==========================
Reads PDF invoices from a watch folder, extracts key data,
fills a weekly Excel report automatically, and renames files.

SETUP — edit the CONFIG section below before running.

Usage: python3 invoice_automation.py
"""

import os
import re
import shutil
from datetime import datetime, timedelta
from pathlib import Path

# ============================================================
#  CONFIG — edit these values for your setup
# ============================================================

# Root folder where your weekly reports live (OneDrive, Dropbox, local, etc.)
BASE_FOLDER = Path.home() / "Documents" / "Invoices"

# Subfolder where new PDFs land (your watch/inbox folder)
INBOX_FOLDER = BASE_FOLDER / "_inbox"

# Your blank Excel template
TEMPLATE_FILE = BASE_FOLDER / "Invoice_Report_Blank.xlsx"

# Company/project name shown in console output
COMPANY_NAME = "My Company"

# Vendors to auto-detect from invoice descriptions.
# Format: { "KEYWORD_IN_PDF": "Label in Excel" }
# Add or remove vendors as needed.
VENDOR_KEYWORDS = {
    "AIRLINE A":  "Airline A",
    "AIRLINE B":  "Airline B",
    "RAIL":       "Rail",
    "CAR RENTAL": "Car Rental",
}

# Fallback vendor label when no keyword matches
DEFAULT_VENDOR = COMPANY_NAME

# Fallback description when no keyword matches
DEFAULT_DESCRIPTION = "Service Fee"

# PDF language keywords (adjust if your invoices are in another language)
KEYWORDS = {
    "passengers":   "Passengers:",       # label before passenger names
    "reference":    "REFERENCE",         # section header for line items
    "payments":     "PAYMENTS",          # section header for payment lines
    "total":        "Total payments",    # end-of-payments marker
    "project":      "Project:",          # project/cost-centre field
    "advisor":      "Advisor:",          # label before invoice date
    "dossier":      "Dossier:",          # file/dossier number
    "skip_desc":    "INDEMNITY FUND",    # description lines to ignore
}

# Excel row range for data (first data row, last data row)
EXCEL_FIRST_ROW = 12
EXCEL_LAST_ROW  = 41

# Excel column positions (1-indexed)
COL_DATE        = 2
COL_VENDOR      = 3
COL_DESCRIPTION = 5
COL_DOSSIER     = 8
COL_INVOICE_NUM = 9
COL_CURRENCY    = 10
COL_AMOUNT      = 12

# ============================================================
#  END CONFIG
# ============================================================

def check_and_install():
    try:
        import pypdf
        import openpyxl
    except ImportError:
        print("Installing required modules...")
        os.system("pip3 install pypdf openpyxl --quiet")

check_and_install()

from pypdf import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

HIGHLIGHT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HIGHLIGHT_FONT = Font(color="0070C0", bold=True)


def get_sunday_of_week():
    today = datetime.today()
    sunday = today + timedelta(days=(6 - today.weekday()))
    return sunday.strftime("%Y%m%d")


def get_or_create_weekly_folder():
    week_name = get_sunday_of_week()
    week_folder = BASE_FOLDER / week_name
    if not week_folder.exists():
        print(f"📁 Creating folder: {week_name}")
        week_folder.mkdir(parents=True)
    else:
        print(f"📁 Folder exists: {week_name}")
    return week_folder


def get_or_create_weekly_excel(week_folder):
    week_name = week_folder.name
    excel_name = f"Invoice_Report_{week_name}.xlsx"
    excel_path = week_folder / excel_name
    if not excel_path.exists():
        if not TEMPLATE_FILE.exists():
            print(f"❌ Template not found: {TEMPLATE_FILE}")
            exit(1)
        shutil.copy2(TEMPLATE_FILE, excel_path)
        print(f"📄 Template copied: {excel_name}")
    else:
        print(f"📄 Excel exists: {excel_name}")
    return excel_path


def detect_vendor(description):
    """
    Returns (vendor_label, needs_review).
    Matches against VENDOR_KEYWORDS defined in CONFIG.
    """
    if not description:
        return "VERIFY", True
    desc_upper = description.upper()
    for keyword, label in VENDOR_KEYWORDS.items():
        if keyword.upper() in desc_upper:
            return label, False
    return "VERIFY", True


def extract_pdf_data(pdf_path):
    """
    Returns a list of charge dicts extracted from the PDF.
    Each dict: date, dossier, vendor, vendor_needs_review,
               description, invoice_num, currency, total,
               passengers, project
    """
    try:
        reader = PdfReader(str(pdf_path))
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"

        lines = text.split('\n')

        dossier    = None
        date       = None
        currency   = None
        passengers = []
        project    = None
        seen_names = set()

        # Dossier number
        m = re.search(rf'{re.escape(KEYWORDS["dossier"])}\s*(\d+)', text)
        if m:
            dossier = m.group(1)

        # Invoice date (line after "Advisor:" label)
        for i, line in enumerate(lines):
            if KEYWORDS["advisor"] in line:
                for j in range(i + 1, min(i + 5, len(lines))):
                    m2 = re.match(r'^\s*(\d{1,2}/\d{1,2}/\d{4})\s*$', lines[j])
                    if m2:
                        try:
                            dt = datetime.strptime(m2.group(1), "%m/%d/%Y")
                            date = dt.strftime("%d-%m-%Y")
                        except Exception:
                            date = m2.group(1)
                        break
                break

        # Project / cost centre
        for line in lines:
            m = re.search(rf'{re.escape(KEYWORDS["project"])}\s*(\S+)', line, re.IGNORECASE)
            if m:
                val = m.group(1).strip()
                if val:
                    project = val
                break

        # Passenger names
        def clean_name(raw):
            raw = re.sub(r'\b(Mrs?|Ms|Mme|M)\.\s*', '', raw, flags=re.IGNORECASE).strip()
            raw = re.split(r'\s+(?:REFUSED?|CDN|PASSPORT)', raw, flags=re.IGNORECASE)[0].strip()
            raw = raw.strip(' -')
            if raw and re.match(r'[A-Za-zÀ-ÿ\-\ ]+$', raw):
                return re.sub(r'\s+', '', raw.title())
            return None

        for i, line in enumerate(lines):
            if KEYWORDS["passengers"] in line:
                m = re.match(rf'(.+?){re.escape(KEYWORDS["passengers"])}', line)
                if m:
                    name = clean_name(m.group(1))
                    if name and name not in seen_names:
                        passengers.append(name)
                        seen_names.add(name)
                for j in range(i + 1, min(i + 5, len(lines))):
                    next_line = lines[j].strip()
                    if re.match(r'(REFERENCE|TOTAL|PAYMENTS|SERVICE FEES)', next_line, re.IGNORECASE):
                        break
                    if next_line:
                        name = clean_name(next_line)
                        if name and name not in seen_names:
                            passengers.append(name)
                            seen_names.add(name)
                break

        # Currency
        if re.search(r'\bUSD\b', text):
            currency = "USD"
        elif re.search(r'\bEUR\b', text):
            currency = "EUR"

        # Payment lines
        payments = []
        in_payments = False
        for line in lines:
            if KEYWORDS["payments"] in line:
                in_payments = True
                continue
            if in_payments:
                if KEYWORDS["total"] in line:
                    break
                m = re.match(r'(\d{8})\s+CAD.*?\s+([-\d,]+\.\d{2})\s*$', line)
                if m:
                    payments.append({
                        'invoice_num': str(int(m.group(1))),
                        'amount':      float(m.group(2).replace(',', ''))
                    })

        # Description lines (skip the ignore keyword from CONFIG)
        descriptions = []
        in_ref       = False
        current_desc = []

        for line in lines:
            if KEYWORDS["reference"] in line:
                in_ref = True
                continue
            if in_ref:
                if KEYWORDS["payments"] in line or re.match(r'\s*Total:', line):
                    if current_desc:
                        desc = ' '.join(current_desc).strip()
                        if not re.search(re.escape(KEYWORDS["skip_desc"]), desc, re.IGNORECASE):
                            desc = re.sub(r'\s+[\d,]+\.\d{2}\s+CAD.*$', '', desc).strip()
                            desc = re.sub(r'\s{2,}', ' ', desc).strip()[:60]
                            descriptions.append(desc)
                    break
                if re.match(r'^\s*[-\d]', line):
                    if current_desc:
                        desc = ' '.join(current_desc).strip()
                        if not re.search(re.escape(KEYWORDS["skip_desc"]), desc, re.IGNORECASE):
                            desc = re.sub(r'\s+[\d,]+\.\d{2}\s+CAD.*$', '', desc).strip()
                            desc = re.sub(r'\s{2,}', ' ', desc).strip()[:60]
                            descriptions.append(desc)
                    current_desc = []
                    continue
                if line.strip():
                    current_desc.append(line.strip())

        # Build one charge per payment line
        charges = []
        used_desc_indices = set()

        for idx, payment in enumerate(payments):
            if idx < len(descriptions):
                di = idx
                while di in used_desc_indices and di < len(descriptions):
                    di += 1
                desc = descriptions[di] if di < len(descriptions) else ""
                if di < len(descriptions):
                    used_desc_indices.add(di)
            else:
                desc = ""

            vendor, needs_review = detect_vendor(desc)

            # If no vendor matched, use default label and description
            if needs_review and not desc:
                vendor       = DEFAULT_VENDOR
                desc         = DEFAULT_DESCRIPTION
                needs_review = False

            charges.append({
                "date":               date,
                "dossier":            dossier,
                "vendor":             vendor,
                "vendor_needs_review": needs_review,
                "description":        desc,
                "invoice_num":        payment["invoice_num"],
                "currency":           currency,
                "total":              payment["amount"],
                "passengers":         passengers,
                "project":            project,
            })

        return charges if charges else None

    except Exception as e:
        print(f"   ⚠️  Error reading PDF: {e}")
        return None


def get_already_processed(ws):
    processed = set()
    for row in ws.iter_rows(
        min_row=EXCEL_FIRST_ROW, max_row=EXCEL_LAST_ROW, values_only=True
    ):
        inv = row[COL_INVOICE_NUM - 1]
        amt = row[COL_AMOUNT - 1]
        if inv and amt:
            processed.add(f"{inv}_{amt}")
    return processed


def fill_excel(excel_path, all_data):
    wb = load_workbook(str(excel_path))
    ws = wb.active
    already_done = get_already_processed(ws)

    added   = 0
    skipped = 0

    for charge_list in all_data:
        if not charge_list:
            continue
        for charge in charge_list:
            key = f"{charge['invoice_num']}_{charge['total']}"
            if key in already_done:
                skipped += 1
                continue

            target_row = None
            for row_num in range(EXCEL_FIRST_ROW, EXCEL_LAST_ROW + 1):
                if (ws.cell(row=row_num, column=COL_DATE).value is None and
                        ws.cell(row=row_num, column=COL_VENDOR).value is None):
                    target_row = row_num
                    break

            if target_row is None:
                print(f"   ⚠️  Excel full — max {EXCEL_LAST_ROW - EXCEL_FIRST_ROW + 1} rows")
                break

            ws.cell(row=target_row, column=COL_DATE).value        = charge["date"]
            ws.cell(row=target_row, column=COL_DESCRIPTION).value = charge["description"]
            ws.cell(row=target_row, column=COL_DOSSIER).value     = charge["dossier"]
            ws.cell(row=target_row, column=COL_INVOICE_NUM).value = charge["invoice_num"]
            ws.cell(row=target_row, column=COL_CURRENCY).value    = charge["currency"]
            ws.cell(row=target_row, column=COL_AMOUNT).value      = charge["total"]

            vendor_cell = ws.cell(row=target_row, column=COL_VENDOR)
            if charge["vendor_needs_review"]:
                vendor_cell.value = "VERIFY"
                vendor_cell.fill  = HIGHLIGHT_FILL
                vendor_cell.font  = HIGHLIGHT_FONT
            else:
                vendor_cell.value = charge["vendor"]

            already_done.add(key)
            added += 1

    wb.save(str(excel_path))
    return added, skipped


def rename_pdf(pdf_path, charge):
    dossier     = charge.get("dossier", "")
    invoice_num = charge.get("invoice_num", "")
    passengers  = charge.get("passengers", [])
    project     = charge.get("project", "")

    if not invoice_num:
        return

    parts    = [f"Invoice_{dossier}", invoice_num]
    parts   += passengers
    if project:
        parts.append(project)

    new_name = "_".join(parts) + ".pdf"
    new_path = pdf_path.parent / new_name

    if pdf_path.name != new_name:
        try:
            pdf_path.rename(new_path)
            print(f"   📝 Renamed: {new_name}")
        except Exception as e:
            print(f"   ⚠️  Could not rename: {e}")
    else:
        print(f"   📝 Name already correct")


def main():
    print("=" * 50)
    print(f"  INVOICE AUTOMATION — {COMPANY_NAME.upper()}")
    print("=" * 50)
    print()

    if not BASE_FOLDER.exists():
        print(f"❌ Base folder not found: {BASE_FOLDER}")
        print("   Update BASE_FOLDER in the CONFIG section.")
        return

    if not INBOX_FOLDER.exists():
        print(f"❌ Inbox folder not found: {INBOX_FOLDER}")
        print("   Update INBOX_FOLDER in the CONFIG section.")
        return

    week_folder = get_or_create_weekly_folder()
    excel_path  = get_or_create_weekly_excel(week_folder)

    pdfs = list(INBOX_FOLDER.glob("*.pdf"))

    if not pdfs:
        print(f"\n📭 No PDFs found in: {INBOX_FOLDER}")
        print("   Add PDFs to the inbox folder and run again.")
        return

    print(f"\n🔍 {len(pdfs)} PDF(s) found\n")

    all_data = []
    for pdf in pdfs:
        print(f"   Reading: {pdf.name}")
        charges = extract_pdf_data(pdf)
        if charges:
            for charge in charges:
                status = "⚠️  VERIFY vendor" if charge["vendor_needs_review"] else "✅"
                currency_label = charge["currency"] or "CAD"
                print(f"   {status} Invoice {charge['invoice_num']} — {charge['vendor']} — {charge['total']} {currency_label}")
            all_data.append(charges)
            rename_pdf(pdf, charges[0])
        else:
            print(f"   ❌ Could not read this PDF")

    print(f"\n📊 Updating Excel...")
    added, skipped = fill_excel(excel_path, all_data)

    print()
    print("=" * 50)
    print(f"  ✅ DONE")
    print(f"  {added} invoice(s) added")
    if skipped > 0:
        print(f"  {skipped} invoice(s) already present — skipped")
    print(f"  File: {excel_path.name}")
    if any(c["vendor_needs_review"] for charges in all_data for c in charges):
        print(f"  ⚠️  Some vendors marked VERIFY in blue — fill in manually")
    print("=" * 50)


if __name__ == "__main__":
    main()
